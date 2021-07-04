"""
@authors: Kfir Ram, Hen Dahan, Shir Bar
@date: 30.6.2021
"""

import unittest
import time
import json
import urllib.request
import os
import sqlite3
import urllib
from urllib.request import urlopen
import re
from openpyxl import load_workbook


def invert_malware_hash_map(mitre_hash_map):
    new_dic = {}
    for k, v in mitre_hash_map.items():
        for x in v:
            if '-' not in x:
                new_dic.setdefault(int(x), []).append(k)
    return new_dic


# This function save the mitre cti tuple to Malware.db file
def save_malware_to_db(malware_hashmap):
    conn = sqlite3.connect("Databases/Malware.db")
    cur = conn.cursor()

    if os.path.exists("Databases/Malware.db"):
        drop = "DROP TABLE IF EXISTS malware"
        cur.execute(drop)

    create = "CREATE TABLE IF NOT EXISTS malware( event_id INT, ttp TEXT);"
    cur.execute(create)  # execute SQL commands
    conn.commit()

    malware_data = malware_hashmap
    malware_data = [(int(i), str(malware_data[i])) for i in malware_data]
    insert_command = "INSERT INTO malware VALUES(?,?);"

    cur.executemany(insert_command, malware_data)
    conn.commit()
    # show_db()


# this function shwos the data inside Malware.db
def show_db():
    conn = sqlite3.connect("Databases/Malware.db")
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table';")  # show all the tables in the .db file
    print("Malware.db__________________________________________________________________________________")
    print(cur.fetchall())
    cur.execute("SELECT * FROM Malware")  # show all the data inside Malware table/
    print(cur.fetchall())
    names = list(map(lambda x: x[0], cur.description))  # show all the columns names
    print(names)
    print("Malware.db_end______________________________________________________________________________")


# This function extract only the event IDs from a string, using a regular expressions library,
# returns a list of event IDs
def extract_event_id_from_str(string):
    return re.findall('[\d]*[-][\d]+|[\d]+[.\d]+', string)


def get_malware_hash_map():
    url = "https://raw.github.com/MalwareArchaeology/ATTACK/6da9ffb897e40e83d3ddd8a995f8af9a0044caca/Windows_Logging_Attack_Matrix_Win_Events_Sept_2018.xlsx"
    urllib.request.urlretrieve(url, 'WindowsMalware.xlsx')

    # Open Malware Archaeology excels sheet file
    workbook = load_workbook(filename="WindowsMalware.xlsx")
    sheet = workbook["Win_Logging_Basic"]

    # initialize sheet values that are necessary for create the hashmap
    rowsNum = sheet.max_row
    columnsNum = sheet.max_column
    Green_Hex = "FF92D050"
    Yellow_Hex = "FFFFFF00"

    # Create a malware HashMap
    malware_hashmap = {}

    # Run on the cells in the sheet for creating the hashmap
    for i in range(6, rowsNum):
        Tactic = sheet.cell(i, 4).value
        for j in range(5, columnsNum):
            cellColor = str(sheet.cell(i, j).fill.start_color.index)
            if cellColor == Green_Hex or cellColor == Yellow_Hex:
                cell = str(sheet.cell(i, j).value)
                eventIDs = extract_event_id_from_str(cell)
                malware_hashmap[Tactic] = eventIDs
            else:
                continue
    # print(malware_hashmap)
    # print(invert_malware_hash_map(malware_hashmap))
    invertHashMap = invert_malware_hash_map(malware_hashmap)
    save_malware_to_db(invertHashMap)
    return invertHashMap


def get_malware_archaeology_hashmap_from_db():
    malware_hashmap = {}
    if not os.path.exists("Databases/Malware.db"):
        get_malware_hash_map()
    try:
        sqliteConnection = sqlite3.connect("Databases/Malware.db")
        cursor = sqliteConnection.cursor()
        sqlite_select_Query = "select event_id, ttp from malware"
        cursor.execute(sqlite_select_Query)
        record = cursor.fetchall()
        for rec in record:
            if rec[0] in malware_hashmap.keys():
                malware_hashmap[int(rec[0])].append(rec[1])
            else:
                # print(rec[0])
                malware_hashmap[int(rec[0])] = [rec[1]]
        cursor.close()
        sqliteConnection.close()
        # print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        # print(malware_hashmap)
        # print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        return malware_hashmap
    except sqlite3.Error as error:
        print("error while connecting to sqlite ", error)


# this function compere the old list from the db with new one from the internet
# return true if update is needed else return false
def check_for_update():
    local_list = get_malware_archaeology_hashmap_from_db()
    foreign_list = get_malware_hash_map()
    for item in local_list:
        if item not in foreign_list:
            return True
    return False


class MyTestCase(unittest.TestCase):
    def test_event_list_update_duration(self):
        t0 = time.time()
        url = "https://raw.githubusercontent.com/miriamxyra/EventList/master/EventList/internal/data/EventList.db"
        urllib.request.urlretrieve(url, '../Databases/EventList.db')
        t1 = time.time()
        duration = t1 - t0
        print("Event List Update Duration: " + str(duration))
        self.assertTrue(True)

    def test_malware_archeology_update_duration(self):
        t0 = time.time()

        url = "https://raw.github.com/MalwareArchaeology/ATTACK/6da9ffb897e40e83d3ddd8a995f8af9a0044caca/Windows_Logging_Attack_Matrix_Win_Events_Sept_2018.xlsx"
        urllib.request.urlretrieve(url, '../WindowsMalware.xlsx')

        # Open Malware Archaeology excels sheet file
        workbook = load_workbook(filename="../WindowsMalware.xlsx")
        sheet = workbook["Win_Logging_Basic"]

        # initialize sheet values that are necessary for create the hashmap
        rowsNum = sheet.max_row
        columnsNum = sheet.max_column
        Green_Hex = "FF92D050"
        Yellow_Hex = "FFFFFF00"

        # Create a malware HashMap
        malware_hashmap = {}

        # Run on the cells in the sheet for creating the hashmap
        for i in range(6, rowsNum):
            Tactic = sheet.cell(i, 4).value
            for j in range(5, columnsNum):
                cellColor = str(sheet.cell(i, j).fill.start_color.index)
                if cellColor == Green_Hex or cellColor == Yellow_Hex:
                    cell = str(sheet.cell(i, j).value)
                    eventIDs = extract_event_id_from_str(cell)
                    malware_hashmap[Tactic] = eventIDs
                else:
                    continue
        # print(malware_hashmap)
        # print(invert_malware_hash_map(malware_hashmap))
        invertHashMap = invert_malware_hash_map(malware_hashmap)
        save_malware_to_db(invertHashMap)

        t1 = time.time()
        duration = t1 - t0
        print("Malware Archeology Update Duration: " + str(duration))
        self.assertTrue(True)

    def test_mitre_cti_update_duration(self):
        def get_mitre_cti_hash_map():
            def load_windows_event_ids():
                try:
                    pattern_file = open("../Util/windows-event-ids.txt")
                    for line in pattern_file:
                        line = line.strip()
                        line_split = line.split('\t', maxsplit=2)
                        event_id_ = line_split[1]
                        header_ = line_split[2]
                        pattern_dict[header_] = event_id_
                    # print(pattern_dict)
                except:
                    print("Error - failed to open the file windows-event-ids.txt")
                    exit()

            def search_pattern(text):
                for header in pattern_dict.keys():
                    if 'x_mitre_detection' in text.keys():
                        for word in str(header).split(' '):
                            if word.lower() not in filter_words and word in text['x_mitre_detection']:
                                technique = text['external_references'][0]['external_id']
                                matched_words.append(word)
                                # print(text['x_mitre_detection'])
                                if (len(technique) > 1) and (technique in mitre_hash_technique.keys()):
                                    mitre_hash_technique[technique].append(pattern_dict[header])
                                else:
                                    mitre_hash_technique[technique] = list(pattern_dict[header])

            pattern_dict = {}
            mitre_hash_technique = {}
            matched_words = []
            filter_words = (
                "a", "or", "to", "the", "for", "an", "its", "and", "of", "in", "from", "was", "on", "it", "has", "have",
                "been", "did", "not", "that", "with", "are", "as", "be", "this", "is", "now", "id", "get", "can", "no",
                "if", "get", "by", "after", "into", "up", "some", "does", "more", "see", "being", "made", "when",
                "only",
                "those", "but", "while", "other", "one", "per", "were", "will", "met", "could", "user", "users", "them",
                "they", "which",)
            load_windows_event_ids()
            url = "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json"
            json_url = urlopen(url)
            data = json.loads(json_url.read())

            for i in data['objects']:
                if 'x_mitre_data_sources' in i:
                    # if 'Windows event logs' in i['x_mitre_data_sources']:
                    if 'Event ID' in i['x_mitre_detection']:
                        eventIds = get_event_ids(i['x_mitre_detection'])
                        key = i['external_references'][0]['external_id']
                        if key in mitre_hash_technique.keys():
                            mitre_hash_technique[key].append(eventIds)
                        else:
                            mitre_hash_technique[key] = eventIds
                search_pattern(i)
            # print(mitre_hash_technique)
            matched_words = set(matched_words)
            print(matched_words)
            print("inverting------")
            return invert_mitre_hash_map(mitre_hash_technique)
            # return mitre_hash_technique

        # This function gets the x_mire_detection str.split string and return the event ids thet can be use with this techniqe
        def get_event_ids(description):
            eventIds = re.findall(r'\b\d+\b', description)
            for event in reversed(eventIds):
                if int(event) < 1100 or 1108 < int(event) < 4608:
                    eventIds.remove(event)
            eventIds = set(eventIds)
            return list(eventIds)

        def invert_mitre_hash_map(mitre_hash_map):
            new_dic = {}
            print("original hash map:")
            print(mitre_hash_map)
            print("-------------\n\n")
            for k, v in mitre_hash_map.items():
                for x in v:
                    if int(x) not in new_dic.keys():
                        new_dic[int(x)] = list(k)
                    elif k not in new_dic[int(x)]:
                        new_dic[int(x)].append(k)

            print(new_dic)
            return new_dic

        def get_lest_modified_date():
            modified_hash_map = {}

            url = "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json"
            json_url = urlopen(url)
            data = json.loads(json_url.read())

            for i in data['objects']:
                if 'modified' in i:
                    if 'external_references' in i:
                        if 'external_id' in i['external_references'][0]:
                            date = i['modified']
                            key = i['external_references'][0]['external_id']
                            if key in modified_hash_map.keys():
                                modified_hash_map[key].append(date)
                            else:
                                modified_hash_map[key] = []
                                modified_hash_map[key].append(date)

                            # print("TTTTTTrue")

            # print(modified_hash_map)
            return modified_hash_map

        t0 = time.time()
        conn = sqlite3.connect("../Databases/Mitre_CTI.db")
        cur = conn.cursor()

        if os.path.exists("../Databases/Mitre_CTI.db"):
            drop = "DROP TABLE IF EXISTS mitre_cti"
            cur.execute(drop)

        create = "CREATE TABLE IF NOT EXISTS mitre_cti( event_id INT, ttp TEXT);"
        cur.execute(create)  # execute SQL commands
        conn.commit()

        mitre_cti_data = get_mitre_cti_hash_map()
        mitre_cti_data = [(int(i), str(mitre_cti_data[i])) for i in mitre_cti_data]
        insert_command = "INSERT INTO mitre_cti VALUES(?,?);"

        cur.executemany(insert_command, mitre_cti_data)
        conn.commit()

        # save to db the lest modify date of the ttps.
        if os.path.exists("../Databases/Mitre_CTI.db"):
            drop = "DROP TABLE IF EXISTS mitre_cti_last_modify"
            cur.execute(drop)

        create = "CREATE TABLE IF NOT EXISTS mitre_cti_last_modify(ttp TEXT, date TEXT);"
        cur.execute(create)  # execute SQL commands
        conn.commit()

        mitre_cti_last_modify = get_lest_modified_date()
        # print_check(mitre_cti_last_modify)
        mitre_cti_last_modify = [(str(i), str(mitre_cti_last_modify[i])) for i in mitre_cti_last_modify]
        # print_check(mitre_cti_last_modify)
        insert_command = "INSERT INTO mitre_cti_last_modify VALUES(?,?);"

        cur.executemany(insert_command, mitre_cti_last_modify)
        conn.commit()
        # show_db()

        t1 = time.time()
        duration = t1 - t0
        print("Mitre Cti Update Duration: " + str(duration))
        self.assertTrue(True)


if __name__ == '__main__':
    unittest.main()
